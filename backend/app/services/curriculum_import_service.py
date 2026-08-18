"""Imports one chapter workbook (the client's own "Portal Practice Bank"
.xlsx format -- see Content/CBSE_Class_5_Chapter_*.xlsx and each file's
"Portal Schema" sheet) into the curriculum + question-bank tables from
app/models/curriculum.py.

New for Phase 2. Column positions below were verified against three
different chapter files (1, 7, 10) before writing this -- the 32-column
"Question Bank" layout is identical in column order across every chapter
(only minor header label wording differs, e.g. "Chapter No" vs "Chapter
No."), so this parses by fixed position rather than by header text match,
which is the more brittle approach given that observed drift.

Import target is always DRAFT status (blueprint Section 15.1: "Import only
into DRAFT state; publishing requires validation and approval") -- this
service never marks anything PUBLISHED itself.

Idempotent by design: every natural code (Board.code, Discipline.code,
BoardCourse.code, Chapter.code, ConceptLesson.code, Question.code) is
looked up before insert, so re-running an import (e.g. after a content
team correction and re-export, blueprint Section 15.1's own "support
export... for correction and re-import" workflow) updates existing rows
rather than duplicating them.
"""
from dataclasses import dataclass, field

import openpyxl
from sqlalchemy.orm import Session

from app.models import (
    Board,
    BoardCourse,
    Chapter,
    ClassLevel,
    ConceptLesson,
    CourseDisciplineMap,
    CurriculumVersion,
    Discipline,
    Question,
    SubjectGroup,
)

# Question Bank sheet column positions (0-indexed) -- see module docstring.
QB_QUESTION_ID = 0
QB_BOARD = 1
QB_CLASS = 2
QB_SUBJECT = 3
QB_CHAPTER_NO = 4
QB_CHAPTER = 5
QB_SKILL_ID = 6
QB_SKILL = 7
QB_ASSIGNMENT_ID = 8
QB_STAGE = 9
QB_DIFFICULTY = 10
QB_COMPETENCY = 11
QB_QUESTION_TYPE = 12
QB_STEM = 13
QB_OPTION_A = 14
QB_OPTION_B = 15
QB_OPTION_C = 16
QB_OPTION_D = 17
QB_CORRECT_ANSWER = 18
QB_ACCEPTED_VARIANTS = 19
QB_HINT = 20
QB_EXPLANATION = 21
QB_MISCONCEPTION_TAG = 22
QB_MARKS = 23
QB_TIME_SECONDS = 24
QB_AUTO_GRADABLE = 25
QB_SHUFFLE_OPTIONS = 26
QB_RESPONSE_FORMAT = 27
QB_MEDIA_REQUIRED = 28
QB_TEACHER_NOTE = 29
QB_STATUS = 30
QB_SOURCE_ALIGNMENT = 31

STATUS_MAP = {
    "draft": "DRAFT",
    "sme review": "SME_REVIEW",
    "approved": "APPROVED",
    "published": "PUBLISHED",
}


@dataclass
class ImportResult:
    board_code: str
    discipline_code: str
    board_course_code: str
    chapter_code: str
    chapter_title: str
    concept_lessons_created: int = 0
    concept_lessons_updated: int = 0
    questions_created: int = 0
    questions_updated: int = 0
    warnings: list = field(default_factory=list)


def _yn_to_bool(value) -> bool:
    return str(value).strip().lower() == "yes"


def _normalize_status(value) -> str:
    if not value:
        return "DRAFT"
    return STATUS_MAP.get(str(value).strip().lower(), "DRAFT")


def _find_header_row(rows, first_cell_value: str) -> int:
    """Skill Map sheets have 1-4 title/subtitle rows before the real header
    (see module docstring's Skill Map dump) -- locate it by content instead
    of assuming a fixed row index, since that varies slightly between files.
    """
    for i, row in enumerate(rows):
        if row and row[0] == first_cell_value:
            return i
    raise ValueError(f"Could not find header row starting with {first_cell_value!r}")


def _get_or_create_board(db: Session, code: str, display_name: str) -> Board:
    board = db.query(Board).filter(Board.code == code).first()
    if board:
        return board
    board = Board(code=code, display_name=display_name)
    db.add(board)
    db.flush()
    return board


def _get_or_create_class_level(db: Session, code: str) -> ClassLevel:
    level = db.query(ClassLevel).filter(ClassLevel.code == code).first()
    if level:
        return level
    level = ClassLevel(code=code, display_name=f"Class {code}", display_order=int(code) if code.isdigit() else 0)
    db.add(level)
    db.flush()
    return level


def _get_or_create_subject_group(db: Session, code: str, display_name: str) -> SubjectGroup:
    group = db.query(SubjectGroup).filter(SubjectGroup.code == code).first()
    if group:
        return group
    group = SubjectGroup(code=code, display_name=display_name)
    db.add(group)
    db.flush()
    return group


def _get_or_create_discipline(db: Session, code: str, display_name: str, subject_group: SubjectGroup) -> Discipline:
    discipline = db.query(Discipline).filter(Discipline.code == code).first()
    if discipline:
        return discipline
    discipline = Discipline(code=code, display_name=display_name, subject_group_id=subject_group.id)
    db.add(discipline)
    db.flush()
    return discipline


def _get_or_create_board_course(
    db: Session, board: Board, class_level: ClassLevel, code: str, display_name: str
) -> BoardCourse:
    course = (
        db.query(BoardCourse)
        .filter(
            BoardCourse.board_id == board.id,
            BoardCourse.class_level_id == class_level.id,
            BoardCourse.code == code,
        )
        .first()
    )
    if course:
        return course
    course = BoardCourse(
        board_id=board.id, class_level_id=class_level.id, code=code, display_name=display_name, status="DRAFT"
    )
    db.add(course)
    db.flush()
    return course


def _get_or_create_curriculum_version(
    db: Session, board: Board, code: str, label: str, status: str, effective_from: str | None
) -> CurriculumVersion:
    version = (
        db.query(CurriculumVersion)
        .filter(CurriculumVersion.board_id == board.id, CurriculumVersion.code == code)
        .first()
    )
    if version:
        return version
    version = CurriculumVersion(
        board_id=board.id, code=code, label=label, status=status, effective_from=effective_from
    )
    db.add(version)
    db.flush()
    return version


def _ensure_course_discipline_map(db: Session, board_course: BoardCourse, discipline: Discipline) -> None:
    existing = (
        db.query(CourseDisciplineMap)
        .filter(
            CourseDisciplineMap.board_course_id == board_course.id,
            CourseDisciplineMap.discipline_id == discipline.id,
        )
        .first()
    )
    if existing:
        return
    db.add(CourseDisciplineMap(board_course_id=board_course.id, discipline_id=discipline.id, sequence=1))
    db.flush()


def import_chapter_workbook(
    db: Session,
    file_path: str,
    *,
    board_code: str = "CBSE",
    board_display_name: str = "CBSE",
    class_level_code: str = "5",
    subject_group_code: str = "SCIENCE",
    subject_group_display_name: str = "Science Group",
    discipline_code: str = "MATHEMATICS",
    discipline_display_name: str = "Mathematics",
    board_course_code: str = "MATHEMATICS",
    board_course_display_name: str = "Mathematics",
    curriculum_version_code: str = "2026-27",
    curriculum_version_label: str = "2026-27",
    curriculum_version_status: str = "PUBLISHED",
    curriculum_version_effective_from: str | None = "2026-04-01",
) -> ImportResult:
    """Import one chapter workbook. Master data (board/class level/subject
    group/discipline/board course/curriculum version) is shared across
    chapters and created once, on first import, then reused -- callers
    importing multiple chapters for the same board/class/subject/edition
    should pass the same codes each time.

    A chapter's real identity is (board_course, discipline, curriculum_version,
    code) -- see the Chapter model docstring (18 Aug 2026) -- not just
    (discipline, code), so that e.g. Class 6 Maths content can safely reuse
    "CH01" the same way Class 5 Maths does, and a future syllabus edition can
    be imported alongside an older one without colliding with it. Callers
    importing a new class or a new syllabus year should pass a distinct
    board_course_code/class_level_code or curriculum_version_code
    accordingly -- reusing the same codes across genuinely different
    classes/editions is what would cause a collision, not this importer.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    qb_rows = list(wb["Question Bank"].iter_rows(values_only=True))
    if not qb_rows or qb_rows[0][QB_QUESTION_ID] != "Question ID":
        raise ValueError("Question Bank sheet does not start with the expected header row")
    qb_data_rows = [r for r in qb_rows[1:] if r and r[QB_QUESTION_ID]]

    sm_rows = list(wb["Skill Map"].iter_rows(values_only=True))
    sm_header_idx = _find_header_row(sm_rows, "Skill ID")
    sm_data_rows = [r for r in sm_rows[sm_header_idx + 1 :] if r and r[0]]

    first = qb_data_rows[0]
    chapter_no = int(first[QB_CHAPTER_NO])
    chapter_title = str(first[QB_CHAPTER]).strip()
    chapter_code = f"CH{chapter_no:02d}"

    board = _get_or_create_board(db, board_code, board_display_name)
    class_level = _get_or_create_class_level(db, class_level_code)
    subject_group = _get_or_create_subject_group(db, subject_group_code, subject_group_display_name)
    discipline = _get_or_create_discipline(db, discipline_code, discipline_display_name, subject_group)
    board_course = _get_or_create_board_course(
        db, board, class_level, board_course_code, board_course_display_name
    )
    _ensure_course_discipline_map(db, board_course, discipline)
    curriculum_version = _get_or_create_curriculum_version(
        db, board, curriculum_version_code, curriculum_version_label,
        curriculum_version_status, curriculum_version_effective_from,
    )

    chapter = (
        db.query(Chapter)
        .filter(
            Chapter.board_course_id == board_course.id,
            Chapter.discipline_id == discipline.id,
            Chapter.curriculum_version_id == curriculum_version.id,
            Chapter.code == chapter_code,
        )
        .first()
    )
    if chapter is None:
        chapter = Chapter(
            discipline_id=discipline.id,
            board_course_id=board_course.id,
            curriculum_version_id=curriculum_version.id,
            code=chapter_code,
            chapter_no=chapter_no,
            title=chapter_title,
            sequence=chapter_no,
            status="DRAFT",
            source_reference=f"Imported from {file_path}",
        )
        db.add(chapter)
        db.flush()
    else:
        chapter.title = chapter_title

    result = ImportResult(
        board_code=board_code,
        discipline_code=discipline_code,
        board_course_code=board_course_code,
        chapter_code=chapter_code,
        chapter_title=chapter_title,
    )

    # --- Concept lessons (one per Skill Map row) ---
    concept_lessons_by_code: dict[str, ConceptLesson] = {}
    for i, row in enumerate(sm_data_rows, start=1):
        skill_id = str(row[0]).strip()
        skill_name = str(row[1]).strip()
        learning_outcome = row[2]
        prerequisite_note = row[3]
        priority_misconception = row[4]

        lesson = (
            db.query(ConceptLesson)
            .filter(ConceptLesson.chapter_id == chapter.id, ConceptLesson.code == skill_id)
            .first()
        )
        if lesson is None:
            lesson = ConceptLesson(
                chapter_id=chapter.id,
                code=skill_id,
                title=skill_name,
                learning_outcome=learning_outcome,
                prerequisite_note=prerequisite_note,
                priority_misconception=priority_misconception,
                sequence=i,
                status="DRAFT",
            )
            db.add(lesson)
            db.flush()
            result.concept_lessons_created += 1
        else:
            lesson.title = skill_name
            lesson.learning_outcome = learning_outcome
            lesson.prerequisite_note = prerequisite_note
            lesson.priority_misconception = priority_misconception
            result.concept_lessons_updated += 1
        concept_lessons_by_code[skill_id] = lesson

    # --- Questions (one per Question Bank row) ---
    for row in qb_data_rows:
        skill_id = str(row[QB_SKILL_ID]).strip()
        lesson = concept_lessons_by_code.get(skill_id)
        if lesson is None:
            result.warnings.append(
                f"Question {row[QB_QUESTION_ID]} references Skill ID {skill_id!r}, not found in Skill Map -- skipped"
            )
            continue

        code = str(row[QB_QUESTION_ID]).strip()
        question = db.query(Question).filter(Question.code == code).first()
        is_new = question is None
        if question is None:
            question = Question(code=code, concept_lesson_id=lesson.id)

        question.concept_lesson_id = lesson.id
        question.assignment_code = row[QB_ASSIGNMENT_ID]
        question.stage = row[QB_STAGE]
        question.difficulty = int(row[QB_DIFFICULTY]) if row[QB_DIFFICULTY] is not None else None
        question.competency = row[QB_COMPETENCY]
        question.question_type = row[QB_QUESTION_TYPE]
        question.stem = row[QB_STEM]
        question.option_a = row[QB_OPTION_A]
        question.option_b = row[QB_OPTION_B]
        question.option_c = row[QB_OPTION_C]
        question.option_d = row[QB_OPTION_D]
        question.correct_answer = str(row[QB_CORRECT_ANSWER]) if row[QB_CORRECT_ANSWER] is not None else ""
        question.accepted_variants = row[QB_ACCEPTED_VARIANTS]
        question.hint = row[QB_HINT]
        question.explanation = row[QB_EXPLANATION]
        question.misconception_tag = row[QB_MISCONCEPTION_TAG]
        question.marks = int(row[QB_MARKS]) if row[QB_MARKS] is not None else 1
        question.time_seconds = int(row[QB_TIME_SECONDS]) if row[QB_TIME_SECONDS] is not None else None
        question.auto_gradable = _yn_to_bool(row[QB_AUTO_GRADABLE])
        question.shuffle_options = _yn_to_bool(row[QB_SHUFFLE_OPTIONS])
        question.response_format = row[QB_RESPONSE_FORMAT]
        question.media_required = row[QB_MEDIA_REQUIRED]
        question.teacher_note = row[QB_TEACHER_NOTE]
        question.status = _normalize_status(row[QB_STATUS])
        question.source_alignment = row[QB_SOURCE_ALIGNMENT]

        if is_new:
            db.add(question)
            result.questions_created += 1
        else:
            result.questions_updated += 1

    db.commit()
    return result
