"""Proves curriculum_import_service.import_chapter_workbook() against a
small synthetic workbook matching the real content team's format (verified
directly against the actual CBSE_Class_5_Chapter_1.xlsx/Chapter_7.xlsx/
Chapter_10.xlsx files during development -- see curriculum_import_service.py
docstring). The real 500-question chapter files are proprietary client
content and are not committed to this public repo, so this fixture is
intentionally small and self-contained rather than checking in real content.

Each test builds a workbook with its own chapter number -- the test DB is an
in-memory SQLite instance shared across the whole session (see conftest.py),
not reset per test, so distinct chapter/question codes avoid cross-test
collisions the same way test_auth.py and test_platform.py already do.
"""
import openpyxl
import pytest

from app.models import Board, BoardCourse, Chapter, ConceptLesson, Question
from app.services.curriculum_import_service import import_chapter_workbook

QUESTION_BANK_HEADER = [
    "Question ID", "Board", "Class", "Subject", "Chapter No", "Chapter",
    "Skill ID", "Skill", "Assignment ID", "Stage", "Difficulty", "Competency",
    "Question Type", "Question Stem", "Option A", "Option B", "Option C", "Option D",
    "Correct Answer", "Accepted Variants", "Hint", "Step-by-Step Explanation",
    "Misconception Tag", "Marks", "Time (seconds)", "Auto-gradable", "Shuffle Options",
    "Response Format", "Media Required", "Teacher Note", "Status", "Source Alignment",
]

SKILL_MAP_HEADER = [
    "Skill ID", "Skill", "Learning Outcome", "Prerequisite / Gap Check",
    "Priority Misconception", "Question Count", "Diagnostic", "Core Practice",
    "Support / Extension", "Coverage Check",
]


def _build_fixture_workbook(path: str, chapter_no: int, title: str) -> None:
    ch = f"CH{chapter_no:02d}"
    skill1, skill2 = f"T5-C{chapter_no}-S01", f"T5-C{chapter_no}-S02"

    wb = openpyxl.Workbook()
    qb = wb.active
    qb.title = "Question Bank"
    qb.append(QUESTION_BANK_HEADER)
    qb.append([
        f"TEST-C5-{ch}-Q001", "CBSE / NCERT", 5, "Mathematics", chapter_no, title,
        skill1, "Test skill one", f"{ch}-DIAG", "Diagnostic", 1, "Fluency",
        "Numeric Entry", "What is 2 + 2?", None, None, None, None,
        "4", "4", "Add the numbers.", "2 + 2 = 4", "Miscounts", 1, 30,
        "Yes", "No", "Integer", "None", None, "Draft", "Test alignment",
    ])
    qb.append([
        f"TEST-C5-{ch}-Q002", "CBSE / NCERT", 5, "Mathematics", chapter_no, title,
        skill2, "Test skill two", f"{ch}-P01", "Core Practice", 2, "Application",
        "Single Select", "Which is even?", "3", "4", "5", "7",
        "B", None, "Even numbers divide by 2.", "4 / 2 = 2", "Confuses odd/even", 1, 45,
        "Yes", "No", "Choice", "None", None, "Published", "Test alignment",
    ])

    sm = wb.create_sheet("Skill Map")
    sm.append([f"{title} Skill Map"])
    sm.append([])
    sm.append(SKILL_MAP_HEADER)
    sm.append([skill1, "Test skill one", "Add small numbers.", "Counting to 10.", "Miscounts", 1, 1, 0, 0, "Complete"])
    sm.append([skill2, "Test skill two", "Identify even numbers.", "Division by 2.", "Confuses odd/even", 1, 0, 1, 0, "Complete"])

    wb.save(path)


@pytest.fixture()
def fixture_workbook(tmp_path):
    path = tmp_path / "test_chapter.xlsx"
    _build_fixture_workbook(str(path), chapter_no=1, title="Test Chapter One")
    return str(path)


def test_import_creates_full_hierarchy(db_session, fixture_workbook):
    result = import_chapter_workbook(db_session, fixture_workbook)

    assert result.chapter_code == "CH01"
    assert result.chapter_title == "Test Chapter One"
    assert result.concept_lessons_created == 2
    assert result.questions_created == 2
    assert result.warnings == []

    board = db_session.query(Board).filter(Board.code == "CBSE").first()
    assert board is not None
    chapter = db_session.query(Chapter).filter(Chapter.code == "CH01").first()
    assert chapter is not None
    assert chapter.status == "DRAFT"

    lessons = db_session.query(ConceptLesson).filter(ConceptLesson.chapter_id == chapter.id).all()
    assert len(lessons) == 2
    assert {l.code for l in lessons} == {"T5-C1-S01", "T5-C1-S02"}

    q1 = db_session.query(Question).filter(Question.code == "TEST-C5-CH01-Q001").first()
    assert q1.question_type == "Numeric Entry"
    assert q1.correct_answer == "4"
    assert q1.status == "DRAFT"  # source said "Draft"

    q2 = db_session.query(Question).filter(Question.code == "TEST-C5-CH01-Q002").first()
    assert q2.status == "PUBLISHED"  # source said "Published" -- normalized to our vocabulary
    assert q2.option_b == "4"


def test_import_is_idempotent(db_session, tmp_path):
    path = tmp_path / "ch2.xlsx"
    _build_fixture_workbook(str(path), chapter_no=2, title="Idempotency Test Chapter")

    import_chapter_workbook(db_session, str(path))
    second = import_chapter_workbook(db_session, str(path))

    assert second.concept_lessons_created == 0
    assert second.concept_lessons_updated == 2
    assert second.questions_created == 0
    assert second.questions_updated == 2

    assert db_session.query(Chapter).filter(Chapter.code == "CH02").count() == 1
    assert db_session.query(Question).filter(Question.code.like("TEST-C5-CH02-%")).count() == 2


def test_second_chapter_reuses_shared_master_data(db_session, tmp_path):
    boards_before = db_session.query(Board).count()
    board_courses_before = db_session.query(BoardCourse).count()

    path_a = tmp_path / "ch3.xlsx"
    _build_fixture_workbook(str(path_a), chapter_no=3, title="Shared Master Data Chapter A")
    import_chapter_workbook(db_session, str(path_a))

    path_b = tmp_path / "ch4.xlsx"
    _build_fixture_workbook(str(path_b), chapter_no=4, title="Shared Master Data Chapter B")
    result = import_chapter_workbook(db_session, str(path_b))
    assert result.chapter_code == "CH04"

    # Master data (board/board_course/discipline) shared, not duplicated --
    # both chapters use the same default CBSE/Mathematics/Class 5 codes.
    assert db_session.query(Board).count() == boards_before
    assert db_session.query(BoardCourse).count() == board_courses_before
    assert db_session.query(Chapter).filter(Chapter.code.in_(["CH03", "CH04"])).count() == 2


def test_import_warns_on_orphan_skill_reference(db_session, tmp_path):
    path = tmp_path / "ch5.xlsx"
    _build_fixture_workbook(str(path), chapter_no=5, title="Orphan Reference Chapter")

    wb = openpyxl.load_workbook(path)
    qb = wb["Question Bank"]
    qb.cell(row=2, column=7, value="T5-C5-S99")  # Skill ID column, not in Skill Map
    wb.save(path)

    result = import_chapter_workbook(db_session, str(path))
    assert len(result.warnings) == 1
    assert "T5-C5-S99" in result.warnings[0]
    assert result.questions_created == 1  # the other, valid row still imports
